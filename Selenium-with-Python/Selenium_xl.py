import openpyxl
import os

# Update path to point to your local .xlsx file
path = os.environ.get("EXCEL_FILE_PATH", "cart.xlsx")
workbook = openpyxl.load_workbook(path)
# For single sheet
sheet = workbook.active
rows = sheet.max_row
col = sheet.max_column

print(rows)
print(col)